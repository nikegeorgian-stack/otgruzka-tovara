# -*- coding: utf-8 -*-
"""Parse FBcell formulation Excel into JSON seed."""
import json
import re
import sys
from pathlib import Path

import openpyxl

XLSX = Path(r"c:\Users\Nika TS\Documents\FBcell\Рецептуры upd. 15.04.2026.xlsx-ის ასლი.xlsx")
OUT = Path(__file__).resolve().parent.parent / "src" / "data" / "seed-formulations.json"

SHEET_CATEGORY = {
    "75": "75",
    "130": "130",
    "145": "145",
    "145 Ультра": "145ultra",
    "160": "160",
    "165 Ультра": "165ultra",
    "МембраныТкань": "membrane",
    "РАТЛ": "ratl",
    "Стеклообои": "glasspaper",
    "Лого": "logo",
}


def slug_code(title: str, sheet: str, idx: int) -> str:
    m = re.search(r"(\d{2,3}[-/][\wА-Яа-яЁё/]+)", title)
    if m:
        raw = re.sub(r"[^\wА-Яа-яЁё/\-]", "", m.group(1))
        return f"РП-{raw}"[:24]
    return f"РП-{sheet}-{idx:03d}"


def parse_recipe_title(title: str) -> dict:
    color = None
    for word, c in [
        ("желт", "yellow"),
        ("оранж", "orange"),
        ("красн", "red"),
        ("син", "blue"),
        ("зелен", "green"),
        ("бел", "white"),
        ("черн", "black"),
        ("сер", "grey"),
    ]:
        if word in title.lower():
            color = c
            break
    variant = None
    m = re.search(r"(\d{2,3}[-/][\wА-Яа-яЁё/]+)", title)
    if m:
        variant = m.group(1).strip()
    grammage = None
    gm = re.search(r"(\d{2,3})\s*г", title)
    if gm:
        grammage = int(gm.group(1))
    return {"colorVariant": color, "variantCode": variant, "grammageGsm": grammage}


def is_recipe_title(v) -> bool:
    if not v or not isinstance(v, str):
        return False
    s = v.lower()
    return "рецептура" in s or "celloplex" in s


def parse_sheet(ws, sheet_name: str) -> list:
    blocks = []
    i = 1
    idx = 0
    while i <= ws.max_row:
        v = ws.cell(i, 1).value
        if not is_recipe_title(v):
            i += 1
            continue
        idx += 1
        title = re.sub(r"\s+", " ", v.replace("\n", " ")).strip()
        hdr = None
        for j in range(i + 1, min(i + 6, ws.max_row + 1)):
            h = ws.cell(j, 1).value
            if h and isinstance(h, str) and "компонент" in h.lower():
                hdr = j
                break
        components = []
        dry_total = None
        batch_total = None
        total_cost = None
        currency = "EUR"
        note_parts = []
        if hdr:
            price_hdr = ws.cell(hdr, 5).value
            if price_hdr and isinstance(price_hdr, str) and "$" in price_hdr:
                currency = "USD"
            for col in range(7, 12):
                n = ws.cell(hdr, col).value or ws.cell(i, col).value
                if n and isinstance(n, str) and n.strip():
                    note_parts.append(n.strip())
            for j in range(hdr + 1, min(hdr + 30, ws.max_row + 1)):
                name = ws.cell(j, 1).value
                if name is None:
                    continue
                if is_recipe_title(name):
                    break
                nm = str(name).strip() if isinstance(name, str) else ""
                if nm and "компонент" in nm.lower():
                    break
                if isinstance(name, str) and nm.lower() in ("итого", ""):
                    dry_total = ws.cell(j, 2).value
                    batch_total = ws.cell(j, 3).value
                    total_cost = ws.cell(j, 6).value
                    break
                w = ws.cell(j, 2).value
                if not nm:
                    continue
                if w is None or isinstance(w, str):
                    continue
                components.append(
                    {
                        "name": nm,
                        "weightKg": float(w) if w is not None else 0,
                        "batchKg": float(ws.cell(j, 3).value)
                        if ws.cell(j, 3).value is not None
                        else None,
                        "sharePct": float(ws.cell(j, 4).value)
                        if ws.cell(j, 4).value is not None
                        else None,
                        "pricePerKg": float(ws.cell(j, 5).value)
                        if ws.cell(j, 5).value is not None
                        else None,
                        "costPerBatch": float(ws.cell(j, 6).value)
                        if ws.cell(j, 6).value is not None
                        else None,
                        "isWater": nm.lower() == "вода",
                    }
                )
        meta = parse_recipe_title(title)
        blocks.append(
            {
                "code": slug_code(title, sheet_name, idx),
                "name": title,
                "sheet": sheet_name,
                "category": SHEET_CATEGORY.get(sheet_name, "other"),
                "variantCode": meta["variantCode"],
                "colorVariant": meta["colorVariant"],
                "grammageGsm": meta["grammageGsm"],
                "currency": currency,
                "dryBatchKg": float(dry_total) if dry_total else None,
                "totalBatchKg": float(batch_total) if batch_total else None,
                "totalCost": float(total_cost) if total_cost else None,
                "note": "\n".join(note_parts) if note_parts else None,
                "components": components,
                "active": True,
            }
        )
        i = i + 1
    return blocks


def parse_pigments(ws) -> list:
    items = []
    for r in range(2, ws.max_row + 1):
        name = ws.cell(r, 1).value
        if not name or not isinstance(name, str):
            continue
        n = name.strip()
        if not n or n.startswith("Kemiteks") or "г/кг" in n or "EUR" in n:
            continue
        ci = ws.cell(r, 2).value
        price = ws.cell(r, 3).value
        if ci is None and price is None:
            continue
        items.append(
            {
                "name": n,
                "colorIndex": str(ci).strip() if ci else None,
                "pricePerKg": float(price) if price is not None else None,
                "currency": "USD",
            }
        )
    return items


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    recipes = []
    for sn in wb.sheetnames:
        if sn in ("Troubleshooting", "Цены", "Пигментные пасты"):
            continue
        recipes.extend(parse_sheet(wb[sn], sn))
    pigments = parse_pigments(wb["Пигментные пасты"])
    data = {"recipes": recipes, "pigmentPastes": pigments}
    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {len(recipes)} recipes, {len(pigments)} pigments -> {OUT}")


if __name__ == "__main__":
    main()
